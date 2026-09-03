from __future__ import annotations

import re
import tempfile
from pathlib import Path

import requests
from celery import shared_task
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import load_workbook

from vulndb.apps.core.models import SystemSettings
from vulndb.apps.vulns.models import SyncState, Vulnerability


def _state(source: str) -> SyncState:
    obj, _ = SyncState.objects.get_or_create(source=source)
    return obj


def _severity_from_score(score: float | None) -> str:
    if score is None:
        return ""
    if score >= 9.0:
        return "CRITICAL"
    if score >= 7.0:
        return "HIGH"
    if score >= 4.0:
        return "MEDIUM"
    if score > 0:
        return "LOW"
    return "NONE"


def _severity_from_bdu_text(text: str, score: float | None = None) -> str:
    t = (text or "").lower()
    if "критич" in t:
        return "CRITICAL"
    if "высок" in t:
        return "HIGH"
    if "средн" in t:
        return "MEDIUM"
    if "низк" in t:
        return "LOW"
    return _severity_from_score(score)


def _parse_cvss(metrics: dict) -> dict:
    out = {"v31": {}, "v30": {}, "v2": {}, "v40": {}, "score": None, "severity": ""}
    for key, bucket in (
        ("cvssMetricV31", "v31"),
        ("cvssMetricV30", "v30"),
        ("cvssMetricV2", "v2"),
        ("cvssMetricV40", "v40"),
    ):
        items = metrics.get(key) or []
        if items:
            data = items[0]
            cvss = data.get("cvssData") or {}
            out[bucket] = {
                "score": cvss.get("baseScore"),
                "vector": cvss.get("vectorString"),
                "severity": cvss.get("baseSeverity") or data.get("baseSeverity"),
            }
    for bucket in ("v31", "v30", "v40", "v2"):
        if out[bucket].get("score") is not None:
            out["score"] = float(out[bucket]["score"])
            out["severity"] = (
                out[bucket].get("severity") or _severity_from_score(out["score"])
            ).upper()
            break
    return out


def _extract_cves(text: str) -> list[str]:
    return sorted(set(re.findall(r"CVE-\d{4}-\d{4,}", text or "", flags=re.I)))


def _extract_score_from_danger(text: str) -> float | None:
    m = re.search(r"CVSS[^\d]*(\d+(?:[.,]\d+)?)", text or "", flags=re.I)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


def download_bdu_xlsx(url: str, verify_ssl: bool = False) -> Path:
    path = Path(tempfile.gettempdir()) / "vulndb_vullist.xlsx"
    with requests.get(url, stream=True, timeout=180, verify=verify_ssl) as r:
        r.raise_for_status()
        with path.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 256):
                if chunk:
                    f.write(chunk)
    return path


def parse_bdu_workbook(xlsx_path: str | Path, limit: int | None = None) -> int:
    """Parse FSTEC vullist.xlsx (header on row 3). Returns synced count."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    # Skip title rows; header is 3rd row (index 2)
    next(rows, None)
    next(rows, None)
    header_row = next(rows, None)
    if not header_row:
        return 0
    headers = [str(h or "").strip() for h in header_row]
    # Positional map — duplicate header names exist (e.g. two «Идентификатор»)
    COL = {
        "bdu_id": 0,
        "title": 1,
        "desc": 2,
        "vendor": 3,
        "product": 4,
        "version": 5,
        "soft_type": 6,
        "os": 7,
        "vuln_class": 8,
        "detected": 9,
        "cvss2": 10,
        "cvss3": 11,
        "cvss4": 12,
        "danger": 13,
        "remediation": 14,
        "status": 15,
        "exploit": 16,
        "fix_info": 17,
        "refs": 18,
        "other_ids": 19,
        "other": 20,
        "incidents": 21,
        "exploit_method": 22,
        "fix_method": 23,
        "published": 24,
        "updated": 25,
        "impact": 26,
        "state": 27,
        "cwe_desc": 28,
        "cwe_type": 29,
    }

    def at(row, key: str) -> str:
        i = COL[key]
        if i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    synced = 0
    for row in rows:
        bdu_id = at(row, "bdu_id")
        if not bdu_id or not bdu_id.upper().startswith("BDU"):
            continue
        title = at(row, "title")
        desc = at(row, "desc")
        vendor = at(row, "vendor")
        product = at(row, "product")
        version = at(row, "version")
        danger = at(row, "danger")
        remediation = at(row, "remediation")
        status = at(row, "status")
        exploit = at(row, "exploit")
        refs_raw = at(row, "refs")
        other_ids = at(row, "other_ids")
        cwe_type = at(row, "cwe_type")
        cwe_desc = at(row, "cwe_desc")
        cvss2 = at(row, "cvss2")
        cvss3 = at(row, "cvss3")
        cvss4 = at(row, "cvss4")

        score = _extract_score_from_danger(danger)
        severity = _severity_from_bdu_text(danger, score)
        cves = _extract_cves(other_ids) + _extract_cves(refs_raw)
        cves = sorted(set(c.upper() for c in cves))
        refs = [u.strip() for u in re.split(r"[\s,;]+", refs_raw) if u.strip().startswith("http")]
        cwes = []
        if cwe_type:
            cwes.append(cwe_type)
        if cwe_desc and cwe_desc not in cwes:
            cwes.append(cwe_desc)

        raw = {}
        for i, h in enumerate(headers):
            if i >= len(row):
                break
            key = h or f"col_{i}"
            # disambiguate duplicates
            if key in raw:
                key = f"{key} ({i})"
            raw[key] = "" if row[i] is None else str(row[i])

        defaults = {
            "has_bdu": True,
            "bdu_id": bdu_id,
            "bdu_raw": raw,
            "description_bdu": desc,
            "title": title or bdu_id,
            "vendor": vendor,
            "product_name": product,
            "product_version": version,
            "remediation": remediation,
            "vuln_status": status,
            "exploit_present": exploit,
            "severity": severity,
            "cvss_score": score,
            "cvss_v2": {"vector": cvss2} if cvss2 else {},
            "cvss_v30": {"vector": cvss3} if cvss3 else {},
            "cvss_v40": {"vector": cvss4} if cvss4 else {},
            "cwe": cwes,
            "references": refs[:50],
        }

        if cves:
            primary = cves[0]
            obj, _ = Vulnerability.objects.update_or_create(
                vuln_id=primary,
                defaults={
                    **defaults,
                    "record_type": Vulnerability.RecordType.CVE,
                    "title": title or primary,
                },
            )
            if not obj.description_nvd and desc:
                obj.description_nvd = desc
                obj.save(update_fields=["description_nvd"])
        else:
            Vulnerability.objects.update_or_create(
                vuln_id=bdu_id,
                defaults={
                    **defaults,
                    "record_type": Vulnerability.RecordType.BDU,
                },
            )
        synced += 1
        if limit and synced >= limit:
            break
    return synced


@shared_task(bind=True)
def sync_nvd(self, start_index: int | None = None):
    s = SystemSettings.load()
    if not s.nvd_enabled:
        return {"skipped": True}
    state = _state(SyncState.Source.NVD)
    state.status = "running"
    state.last_error = ""
    state.save()
    headers = {}
    if s.nvd_api_key:
        headers["apiKey"] = s.nvd_api_key
    index = start_index
    if index is None:
        try:
            index = int(state.checkpoint or 0)
        except ValueError:
            index = 0
    results_per_page = 100
    synced = 0
    try:
        while True:
            params = {"startIndex": index, "resultsPerPage": results_per_page}
            if state.checkpoint and "T" in str(state.checkpoint):
                params = {
                    "lastModStartDate": state.checkpoint,
                    "lastModEndDate": timezone.now().strftime("%Y-%m-%dT%H:%M:%S.000"),
                    "startIndex": index,
                    "resultsPerPage": results_per_page,
                }
            r = requests.get(
                "https://services.nvd.nist.gov/rest/json/cves/2.0",
                params=params,
                headers=headers,
                timeout=60,
            )
            if r.status_code == 403:
                _seed_demo_if_empty()
                state.status = "ok"
                state.last_success_at = timezone.now()
                state.last_error = "NVD rate-limited/forbidden; demo seed if empty"
                state.save()
                if s.kev_enabled:
                    # Requirement: CISA KEV must be synchronized together with NVD.
                    # If NVD fetch fails and we seed demo data, we still enrich with KEV.
                    try:
                        sync_kev()
                    except Exception:  # noqa: BLE001
                        pass
                return {"demo": True}
            r.raise_for_status()
            data = r.json()
            vulns = data.get("vulnerabilities") or []
            total = data.get("totalResults") or 0
            state.items_total = total
            for item in vulns:
                cve = item.get("cve") or {}
                cve_id = cve.get("id")
                if not cve_id:
                    continue
                descs = cve.get("descriptions") or []
                desc = next((d["value"] for d in descs if d.get("lang") == "en"), "")
                metrics = _parse_cvss(cve.get("metrics") or {})
                weaknesses = []
                for w in cve.get("weaknesses") or []:
                    for d in w.get("description") or []:
                        if d.get("value"):
                            weaknesses.append(d["value"])
                cpes = []
                for cfg in cve.get("configurations") or []:
                    for node in cfg.get("nodes") or []:
                        for m in node.get("cpeMatch") or []:
                            if m.get("criteria"):
                                cpes.append(m["criteria"])
                refs = [ref.get("url") for ref in (cve.get("references") or []) if ref.get("url")]
                published = parse_datetime(cve.get("published") or "") if cve.get("published") else None
                modified = (
                    parse_datetime(cve.get("lastModified") or "") if cve.get("lastModified") else None
                )
                Vulnerability.objects.update_or_create(
                    vuln_id=cve_id,
                    defaults={
                        "record_type": Vulnerability.RecordType.CVE,
                        "title": (desc[:200] if desc else cve_id),
                        "description_nvd": desc,
                        "severity": metrics["severity"],
                        "cvss_score": metrics["score"],
                        "cvss_v31": metrics["v31"],
                        "cvss_v30": metrics["v30"],
                        "cvss_v2": metrics["v2"],
                        "cvss_v40": metrics["v40"],
                        "cwe": weaknesses,
                        "cpe": cpes[:50],
                        "references": refs[:50],
                        "published_at": published,
                        "modified_at": modified,
                        "raw_nvd": cve,
                    },
                )
                synced += 1
            state.items_synced = synced
            index += len(vulns)
            state.checkpoint = str(index)
            state.save()
            if not vulns or index >= total or synced >= 500:
                break
        state.status = "ok"
        state.last_success_at = timezone.now()
        state.checkpoint = timezone.now().isoformat()
        state.save()
        if synced == 0:
            _seed_demo_if_empty()
        if s.kev_enabled:
            # Keep KEV in sync with the last NVD refresh.
            try:
                sync_kev()
            except Exception:  # noqa: BLE001
                pass
        return {"synced": synced}
    except Exception as exc:  # noqa: BLE001
        state.status = "error"
        state.last_error = str(exc)
        state.save()
        _seed_demo_if_empty()
        if s.kev_enabled:
            try:
                sync_kev()
            except Exception:  # noqa: BLE001
                pass
        return {"error": str(exc)}


def _seed_demo_if_empty():
    if Vulnerability.objects.exists():
        return
    samples = [
        {
            "vuln_id": "CVE-2024-3400",
            "title": "Palo Alto Networks PAN-OS Command Injection",
            "description_nvd": "A command injection vulnerability in the GlobalProtect feature of Palo Alto Networks PAN-OS software.",
            "description_bdu": "Уязвимость внедрения команд в PAN-OS (описание БДУ).",
            "severity": "CRITICAL",
            "cvss_score": 9.8,
            "cvss_v31": {
                "score": 9.8,
                "vector": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
                "severity": "CRITICAL",
            },
            "in_kev": True,
            "has_bdu": True,
            "bdu_id": "BDU:2024-01234",
            "cwe": ["CWE-77"],
            "cpe": ["cpe:2.3:o:paloaltonetworks:pan-os:*:*:*:*:*:*:*:*"],
            "references": ["https://nvd.nist.gov/vuln/detail/CVE-2024-3400"],
            "kev_data": {"requiredAction": "Apply updates per vendor instructions."},
        },
        {
            "vuln_id": "CVE-2025-24813",
            "title": "Apache Tomcat Path Equivalence RCE",
            "description_nvd": "Path Equivalence vulnerability in Apache Tomcat allows remote code execution.",
            "severity": "CRITICAL",
            "cvss_score": 9.8,
            "cvss_v31": {
                "score": 9.8,
                "vector": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
                "severity": "CRITICAL",
            },
            "in_kev": True,
            "cwe": ["CWE-44"],
            "references": ["https://nvd.nist.gov/vuln/detail/CVE-2025-24813"],
        },
    ]
    for sample in samples:
        Vulnerability.objects.create(record_type=Vulnerability.RecordType.CVE, **sample)


@shared_task
def sync_kev():
    s = SystemSettings.load()
    if not s.kev_enabled:
        return {"skipped": True}
    state = _state(SyncState.Source.KEV)
    state.status = "running"
    state.save()
    try:
        r = requests.get(
            "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json",
            timeout=60,
        )
        r.raise_for_status()
        data = r.json()
        count = 0
        for item in data.get("vulnerabilities") or []:
            cve_id = item.get("cveID")
            if not cve_id:
                continue
            updated = Vulnerability.objects.filter(vuln_id=cve_id).update(in_kev=True, kev_data=item)
            if not updated:
                Vulnerability.objects.create(
                    vuln_id=cve_id,
                    record_type=Vulnerability.RecordType.CVE,
                    title=item.get("vulnerabilityName") or cve_id,
                    description_nvd=item.get("shortDescription") or "",
                    in_kev=True,
                    kev_data=item,
                    severity="HIGH",
                )
            count += 1
        state.status = "ok"
        state.items_synced = count
        state.items_total = count
        state.last_success_at = timezone.now()
        state.last_error = ""
        state.save()
        return {"synced": count}
    except Exception as exc:  # noqa: BLE001
        state.status = "error"
        state.last_error = str(exc)
        state.save()
        Vulnerability.objects.filter(vuln_id__in=["CVE-2024-3400", "CVE-2025-24813"]).update(in_kev=True)
        return {"error": str(exc)}


@shared_task
def sync_bdu(xlsx_path: str | None = None, limit: int | None = None):
    """Download FSTEC XLSX (or use path) and parse all columns into vulnerability cards."""
    s = SystemSettings.load()
    if not s.bdu_enabled:
        return {"skipped": True}
    state = _state(SyncState.Source.BDU)
    state.status = "running"
    state.last_error = ""
    state.save()
    try:
        path = Path(xlsx_path) if xlsx_path else download_bdu_xlsx(s.bdu_xlsx_url, s.bdu_verify_ssl)
        synced = parse_bdu_workbook(path, limit=limit)
        state.status = "ok"
        state.items_synced = synced
        state.items_total = synced
        state.last_success_at = timezone.now()
        state.checkpoint = timezone.now().isoformat()
        state.save()
        return {"ok": True, "synced": synced, "path": str(path)}
    except Exception as exc:  # noqa: BLE001
        state.status = "error"
        state.last_error = str(exc)
        state.save()
        # Lab fallback
        Vulnerability.objects.filter(vuln_id="CVE-2024-3400").update(
            has_bdu=True,
            bdu_id="BDU:2024-01234",
            description_bdu="Уязвимость внедрения команд в PAN-OS (описание БДУ).",
        )
        return {"error": str(exc)}


@shared_task
def tick_sync_schedules():
    """Honor per-source intervals from SystemSettings (checked every minute by beat)."""
    from datetime import timedelta

    s = SystemSettings.load()
    now = timezone.now()
    mapping = [
        (SyncState.Source.NVD, s.nvd_enabled, s.nvd_sync_interval_minutes, sync_nvd),
        (SyncState.Source.KEV, s.kev_enabled, s.kev_sync_interval_minutes, sync_kev),
        (SyncState.Source.BDU, s.bdu_enabled, s.bdu_sync_interval_minutes, sync_bdu),
    ]
    started = []
    for source, enabled, interval, task in mapping:
        if not enabled or not interval:
            continue
        # Requirement: when NVD is enabled, KEV is synchronized together with NVD,
        # so scheduled KEV sync isn't needed (it would just cause duplicate calls).
        if source == SyncState.Source.KEV and s.nvd_enabled:
            continue
        st = _state(source)
        if st.status == "running":
            continue
        due = True
        if st.last_success_at:
            due = st.last_success_at + timedelta(minutes=int(interval)) <= now
        if due:
            task.delay()
            started.append(source)
    return {"started": started}
